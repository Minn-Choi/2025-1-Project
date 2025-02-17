import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from tkinter import Tk, Toplevel, ttk, Label, Entry, Button, filedialog, Checkbutton, BooleanVar, Listbox, END, Frame, StringVar
import datetime
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sys
from tkinter import Toplevel, Button
from tkinter.ttk import Treeview, Scrollbar, Style

plt.rc('font', family='Malgun Gothic')  
plt.rcParams['axes.unicode_minus'] = False 

def select_file():
    """사용자 파일 선택"""
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="엑셀 파일 선택",
        filetypes=[("Excel Files", "*.xls;*.xlsx")]
    )
    return file_path

def get_sorting_order(departments, positions):
    """정렬 순서를 선택하거나 사용자 지정으로 입력받는 함수"""
    root = Toplevel()
    root.title("정렬 순서 선택")
    root.geometry("500x500") 

    departments = ["안전감사실", "임원", "전략기획실", "경영지원부", "체육사업부","주차사업부","시설관리부","사회서비스단"] 
    positions = ["임원", "2급", "3급","4급","5급","6급","7급","시설안내원","주차관리원","지도직","환경관리","비서","사무보조"] 

    use_custom_order = BooleanVar()
    use_custom_order.set(True)

    Label(root, text="정렬 방식을 선택하세요", font=("Arial", 14)).pack(pady=10)
    Checkbutton(root, text="사용자 지정 정렬", variable=use_custom_order).pack()

    custom_order_frame = Frame(root)
    custom_order_frame.pack(fill="both", expand=True, padx=10, pady=10)

    Label(custom_order_frame, text="부서 순서").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    department_listbox = Listbox(custom_order_frame, selectmode="extended", exportselection=False, height=10)
    for department in departments:
        department_listbox.insert(END, department)
    department_listbox.grid(row=1, column=0, padx=5, pady=5)

    Label(custom_order_frame, text="직급 순서").grid(row=0, column=1, padx=5, pady=5, sticky="w")
    position_listbox = Listbox(custom_order_frame, selectmode="extended", exportselection=False, height=10)
    for position in positions:
        position_listbox.insert(END, position)
    position_listbox.grid(row=1, column=1, padx=5, pady=5)

    def move_up(listbox):
        selection = listbox.curselection()
        if not selection or selection[0] == 0:
            return
        for index in selection:
            value = listbox.get(index)
            listbox.delete(index)
            listbox.insert(index - 1, value)
            listbox.selection_set(index - 1)

    def move_down(listbox):
        selection = listbox.curselection()
        if not selection or selection[-1] == listbox.size() - 1:
            return
        for index in reversed(selection):
            value = listbox.get(index)
            listbox.delete(index)
            listbox.insert(index + 1, value)
            listbox.selection_set(index + 1)

    Button(custom_order_frame, text="⬆️ 위로", command=lambda: move_up(department_listbox)).grid(row=2, column=0, pady=5)
    Button(custom_order_frame, text="⬇️ 아래로", command=lambda: move_down(department_listbox)).grid(row=3, column=0, pady=5)

    Button(custom_order_frame, text="⬆️ 위로", command=lambda: move_up(position_listbox)).grid(row=2, column=1, pady=5)
    Button(custom_order_frame, text="⬇️ 아래로", command=lambda: move_down(position_listbox)).grid(row=3, column=1, pady=5)

    custom_order = {"department": [], "position": []}

    def submit():
        if use_custom_order.get():
            custom_order["department"] = [department_listbox.get(i) for i in range(department_listbox.size())]
            custom_order["position"] = [position_listbox.get(i) for i in range(position_listbox.size())]
        root.quit() 
        root.destroy() 

    Button(root, text="확인", command=submit).pack(pady=10)

    root.mainloop()
    return use_custom_order.get(), custom_order

def get_quota_input(departments, positions):
    """부서 및 직급별 정원을 한 화면에서 입력받거나 파일에서 불러오는 기능 추가"""
    root = Toplevel()
    root.title("정원 입력")

    entries = {}
    row = 0

    Label(root, text="정원을 입력하세요", font=("Arial", 14)).grid(row=row, column=0, columnspan=3 + len(positions), pady=10)
    row += 1

    Label(root, text="부서").grid(row=row, column=0, padx=5, pady=5)
    for col, position in enumerate(positions, start=1):
        Label(root, text=position).grid(row=row, column=col, padx=5, pady=5)
    row += 1

    for department in departments:
        Label(root, text=department).grid(row=row, column=0, padx=5, pady=5, sticky="e")
        for col, position in enumerate(positions, start=1):
            entry = Entry(root, width=5)
            entry.grid(row=row, column=col, padx=5, pady=8)
            entries[(department, position)] = entry
        row += 1

    detailed_quota = {}

    def load_from_file():
        """파일에서 정원을 불러오는 기능"""
        file_path = filedialog.askopenfilename(
            title="정원 파일 선택",
            filetypes=[("Excel Files", "*.xls;*.xlsx")]
        )
        if not file_path:
            return
        try:
            quota_df = pd.read_excel(file_path)
            for department, position in entries.keys():
                if (department, position) in quota_df.set_index(["부서", "직급"]).index:
                    value = quota_df.loc[
                        (quota_df["부서"] == department) & (quota_df["직급"] == position),
                        "정원"
                    ].values[0]
                    entries[(department, position)].delete(0, END)
                    entries[(department, position)].insert(0, str(value))
        except Exception as e:
            print(f"파일 로드 오류: {e}")

    def save_to_excel():
        """입력된 정원을 엑셀 파일로 저장"""
        data = []
        for (department, position), entry in entries.items():
            value = entry.get()
            data.append({
                "부서": department,
                "직급": position,
                "정원": int(value) if value.isdigit() else 0
            })
        df = pd.DataFrame(data)
        file_name = f"정원_입력_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        df.to_excel(file_name, index=False)
        print(f"✅ 정원 입력 내용이 저장되었습니다: {file_name}")

    def submit():
        """정원 입력 완료"""
        nonlocal detailed_quota
        for (department, position), entry in entries.items():
            value = entry.get()
            if department not in detailed_quota:
                detailed_quota[department] = {}
            detailed_quota[department][position] = int(value) if value.isdigit() else 0
        save_to_excel()
        root.destroy()

    Button(root, text="파일 선택", command=load_from_file).grid(row=row, column=0, columnspan=1, pady=10)
    Button(root, text="입력 완료", command=submit).grid(row=row, column=1, columnspan=3 + len(positions), pady=10)
    root.wait_window()
    return detailed_quota

def preprocess_data(input_df):
    """데이터 전처리"""
    department_mapping = {
        "85": "임원",
        "84": "안전감사실",
        "81": "전략기획실",
        "82": "경영지원부",
        "83": "체육사업부",
        "86": "주차사업부",
        "91": "시설관리부",
        "89": "사회서비스단",
    }

    sub_department_mapping = {
        "8100": "전략기획부_부장",
        "8101": "기획조정팀",
        "8103": "소통경영팀",
        "8200": "경영지원부_부장",
        "8201": "인사총무팀",
        "8202": "재정지원팀",
        "8300": "체육사업부_부장",
        "8301": "체육사업1팀",
        "8302": "체육사업2팀",
        "8352": "충무",
        "8351": "무학봉",
        "8358": "훈련원공원",
        "8354": "회현",
        "8353": "손기정",
        "8361": "남산",
        "8360": "장충",
        "8400": "안전감사실_부장",
        "8401": "청렴감사팀",
        "8402": "안전보건팀",
        "8500": "임원",
        "8600": "주차사업부_부장",
        "8601": "주차사업1팀",
        "8602": "주차사업2팀",
        "8900": "사회서비스단",
        "9100": "시설관리부_부장",
        "9101": "시설관리1팀",
        "9102": "시설관리2팀",
        "9103": "공공시설팀",
        "9155": "중구구민회관",
        "9157": "공중공원화장실",
        "9152": "중구종합복지센터",
        "9153": "신당누리센터",
        "9154": "교육지원센터",
        "9151": "중립종합복지센터",
        "9156": "충무창업큐브",
    }

    input_df["부서.1"] = input_df["부서"].astype(str).str[:2].map(department_mapping)

    input_df["세부부서"] = input_df["부서"].astype(str).map(sub_department_mapping).fillna("기타")

    input_df["직급"] = input_df["직급"].replace(
        {"이사장": "임원", "본부장": "임원", "수영강사": "지도직", "헬스강사": "지도직", "테니스강사": "지도직", "환경관리원":"환경관리"}
    )

    periodic_workers = input_df[input_df["직종"] == "기간제근로자"].copy()

    input_df = input_df[input_df["직종"] != "기간제근로자"]

    input_df[["부서.1", "세부부서", "직급"]] = input_df[["부서.1", "세부부서", "직급"]].fillna(method="ffill")

    input_df.fillna("", inplace=True)

    return input_df, periodic_workers

def save_to_excel_and_continue(input_df, periodic_workers, original_file_path):
    try:
        if original_file_path.endswith(".xls"):
            save_path = original_file_path.replace(
                ".xls", f"_수정본_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
        elif original_file_path.endswith(".xlsx"):
            save_path = original_file_path.replace(
                ".xlsx", f"_수정본_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
        else:
            raise ValueError(f"지원하지 않는 파일 확장자: {original_file_path}")

        full_df = pd.concat([input_df, periodic_workers], ignore_index=True)  

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            full_df.to_excel(writer, sheet_name="전체직원", index=False)
            periodic_workers.to_excel(writer, sheet_name="기간제근로자", index=False) 

        print(f"✅ 수정된 데이터가 엑셀 파일로 저장되었습니다: {save_path}")

    except Exception as e:
        print(f"❌ 엑셀 파일 저장 중 오류 발생: {e}")


def detect_misclassified_employees(input_df):
    valid_positions = ["임원", "2급", "3급", "4급"]
    invalid_positions = ["5급", "6급", "7급", "시설안내원", "주차관리원", "지도직", "환경관리", "비서", "사무보조"]
    current_year = datetime.datetime.now().year

    def is_misclassified(row):
        birth_date = pd.to_datetime(row.get("생년월일"), errors="coerce")
        age = None
        if not pd.isna(birth_date):
            age = current_year - birth_date.year

        is_main_department = str(row.get("부서", "")).endswith("00")
        position = row.get("직급", "")

        if is_main_department:
            if position in valid_positions and (age is not None and age > 57):
                return True
            elif position in invalid_positions:
                return True
        return False

    misclassified_df = input_df[input_df.apply(is_misclassified, axis=1)]
    columns_to_display = ["사원번호", "한글명", "부서", "세부부서", "직급", "생년월일"]
    missing_columns = [col for col in columns_to_display if col not in input_df.columns]
    if missing_columns:
        print(f"다음 열이 누락되었습니다: {missing_columns}")
        return pd.DataFrame()
    return misclassified_df[columns_to_display] 

def show_misclassified_employees_ui(misclassified_df, input_df, original_file_path, periodic_workers):
    if misclassified_df.empty:
        print("잘못 분류된 데이터가 없습니다.")
        return

    sub_department_mapping = {
        "8100": "전략기획부_부장",
        "8101": "기획조정팀",
        "8103": "소통경영팀",
        "8200": "경영지원부_부장",
        "8201": "인사총무팀",
        "8202": "재정지원팀",
        "8300": "체육사업부_부장",
        "8301": "체육사업1팀",
        "8302": "체육사업2팀",
        "8352": "충무",
        "8351": "무학봉",
        "8358": "훈련원공원",
        "8354": "회현",
        "8353": "손기정",
        "8361": "남산",
        "8360": "장충",
        "8400": "안전감사실_부장",
        "8401": "청렴감사팀",
        "8402": "안전보건팀",
        "8500": "임원",
        "8600": "주차사업부_부장",
        "8601": "주차사업1팀",
        "8602": "주차사업2팀",
        "8900": "사회서비스단",
        "9100": "시설관리부_부장",
        "9101": "시설관리1팀",
        "9102": "시설관리2팀",
        "9103": "공공시설팀",
        "9155": "중구구민회관",
        "9157": "공중공원화장실",
        "9152": "중구종합복지센터",
        "9153": "신당누리센터",
        "9154": "교육지원센터",
        "9151": "중립종합복지센터",
        "9156": "충무창업큐브",
    }

    root = Toplevel()
    root.title("잘못 분류된 데이터 수정")
    root.geometry("800x600")

    Label(root, text="잘못 분류된 데이터를 수정하세요", font=("Arial", 14)).pack(pady=10)

    tree = Treeview(root, columns=list(misclassified_df.columns), show="headings", height=15, selectmode="extended")

    for col in misclassified_df.columns:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="center")

    for _, row in misclassified_df.iterrows():
        tree.insert("", "end", values=list(row))

    tree.pack(fill="both", expand=True, padx=10, pady=10)

    input_frame = Frame(root)
    input_frame.pack(pady=10)

    selected_code = StringVar()
    selected_code.set("8100")

    Label(input_frame, text="새로운 부서 선택").grid(row=0, column=0, padx=5, pady=5)
    dropdown = ttk.Combobox(input_frame, textvariable=selected_code, state="readonly", width=25)
    dropdown["values"] = [f"{code}({name})" for code, name in sub_department_mapping.items()]
    dropdown.grid(row=0, column=1, padx=5, pady=5)

    def apply_changes():
        selected_items = tree.selection() 
        if not selected_items:
            print("수정할 항목을 선택해주세요.")
            return

        selected_value = selected_code.get()
        new_code, new_name = selected_value.split("(")[0], selected_value.split("(")[1][:-1]

        for item in selected_items:
            values = tree.item(item, "values")
            if not values:
                continue

            name = values[1]  
            if name in input_df["한글명"].values:
                input_df.loc[input_df["한글명"] == name, "부서"] = new_code
                input_df.loc[input_df["한글명"] == name, "세부부서"] = new_name

                updated_values = list(values)
                updated_values[2] = new_code  
                updated_values[3] = new_name  
                tree.item(item, values=tuple(updated_values))

                print(f"'{name}'의 부서가 '{new_code}({new_name})'로 수정되었습니다.")

        print(f"✅ {len(selected_items)}명의 부서가 변경되었습니다.")

    def on_done():
        try:
            root.destroy()
            save_to_excel_and_continue(input_df, periodic_workers, original_file_path)

            main(skip_misclassified_check=True)

        except Exception as e:
            print(f"❌ 저장 후 재시작 중 오류 발생: {e}")

    Button(input_frame, text="수정 적용", command=apply_changes).grid(row=1, column=0, columnspan=2, pady=10)
    Button(root, text="수정 완료 및 저장", command=on_done).pack(side="left", padx=10, pady=10)

    root.mainloop()

def traverse_hierarchy_and_count(hierarchy, input_df):
    updated_hierarchy = {}

    for code, data in hierarchy.items():
        sub_departments = data["sub_departments"]
        current_count = input_df[input_df["부서"].astype(str) == code].shape[0]
        sub_department_data = traverse_hierarchy_and_count(sub_departments, input_df)
        sub_total_count = sum(sub_data["현원"] for sub_data in sub_department_data.values())
        total_count = current_count + sub_total_count
        updated_hierarchy[code] = {
            "name": data["name"],
            "현원": total_count,
            "sub_departments": sub_department_data
        }
    return updated_hierarchy

def write_hierarchy_to_excel(ws, hierarchy, input_df, parent=None, color_cycle=None, is_top_level=True):
    if color_cycle is None:
        color_cycle = ["FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "CCFFFF", "FFCCFF"]

    current_color_index = 0 
    for code, data in hierarchy.items():
        name = data["name"]
        current_count = data["현원"]
        sub_departments = data["sub_departments"]

        names = input_df[input_df["부서"].astype(str) == code]["한글명"].tolist()

        row = ws.max_row + 1

        ws.cell(row=row, column=1, value=name) 
        ws.cell(row=row, column=2, value=current_count) 

        if is_top_level:
            fill = PatternFill(start_color=color_cycle[current_color_index % len(color_cycle)],
                               end_color=color_cycle[current_color_index % len(color_cycle)],
                               fill_type="solid")
            ws.cell(row=row, column=1).fill = fill 
            ws.cell(row=row, column=2).fill = fill
            current_color_index += 1 

        if names:
            for i, person_name in enumerate(names):
                ws.cell(row=row + i + 1, column=3, value=person_name) 

        write_hierarchy_to_excel(ws, sub_departments, input_df, parent=None, color_cycle=color_cycle, is_top_level=False)

def create_excel_file(
    department_counts,
    position_counts,
    grouped_counts,
    total_quota,
    position_quota,
    department_quota,
    detailed_quota,
    sub_department_counts,
    department_hierarchy,
    input_df,
    department_order,
    position_order,
    periodic_workers    
):
    """엑셀 파일 생성 및 데이터 기록"""

    for department, positions in detailed_quota.items():
        department_quota[department] = sum(positions.values())

    for position in position_order:
        position_quota[position] = sum(
            detailed_quota[department].get(position, 0) for department in detailed_quota
        )

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "정원 및 현황"
    
    def calculate_dates(row):
        birth_date = row.get("생년월일")

        if pd.isna(birth_date) or birth_date is None:
            return None, None, None, None  

        try:
            birth_date = pd.to_datetime(birth_date, errors="coerce")  
            if pd.isna(birth_date):
                return None, None, None, None

            birth_year = birth_date.year
            birth_month_day = birth_date.strftime("%m-%d")  

            if "01-01" <= birth_month_day <= "06-30":
                retirement_date = f"{birth_year + 60}-06-30"
                base_year = birth_year + 60
                wage_peak_1 = f"{base_year - 3}-07-01"
                wage_peak_2 = f"{base_year - 2}-07-01"
                wage_peak_3 = f"{base_year - 1}-07-01"
            elif "07-01" <= birth_month_day <= "12-31": 
                retirement_date = f"{birth_year + 60}-12-31"
                base_year = birth_year + 60
                wage_peak_1 = f"{base_year - 2}-01-01"
                wage_peak_2 = f"{base_year - 1}-01-01"
                wage_peak_3 = f"{base_year}-01-01"
            else:
                return None, None, None, None

            return wage_peak_1, wage_peak_2, wage_peak_3, retirement_date

        except Exception as e:
            print(f"⚠️ 생년월일 처리 오류: {e}")
            return None, None, None, None  

    input_df[[ "임금피크(1)", "임금피크(2)", "임금피크(3)", "정년"]] = input_df.apply(
        lambda row: pd.Series(calculate_dates(row)), axis=1
    )

    filtered_df = input_df[input_df["생년월일"].notna()]

    ws_all_employees = wb.create_sheet(title="정년 및 임금피크")
    header = ["이름", "직급", "입사일", "부서", "생년월일", "임금피크(1)", "임금피크(2)", "임금피크(3)", "정년"]
    ws_all_employees.append(header)

    header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  
    header_font = Font(bold=True)  
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )  

    for col_num, column_title in enumerate(header, start=1):
        cell = ws_all_employees.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(filtered_df.iterrows(), start=2):  
        ws_all_employees.append([  
            row["한글명"], 
            row["직급"], 
            row["입사일"],
            row["세부부서"], 
            row["생년월일"],
            row["임금피크(1)"], 
            row["임금피크(2)"], 
            row["임금피크(3)"], 
            row["정년"]  
        ])

    date_columns = [5, 6, 7, 8, 9] 

    for col in date_columns:
        col_letter = get_column_letter(col)
        for row_idx in range(2, ws_all_employees.max_row + 1):
            cell = ws_all_employees[f"{col_letter}{row_idx}"]
            if cell.value: 
                try:
                    cell.value = pd.to_datetime(cell.value).date() 
                    cell.number_format = "YYYY-MM-DD" 
                except Exception as e:
                    print(f"⚠️ 날짜 형식 변환 오류 (행 {row_idx}, 열 {col_letter}): {e}")

    ws_all_employees.auto_filter.ref = f"A1:I{ws_all_employees.max_row}"

    for row in ws_all_employees.iter_rows(min_row=1, max_row=ws_all_employees.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.border = thin_border 

    light_pink_fill = PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid")
    for row in ws_all_employees.iter_rows(min_row=2, max_row=ws_all_employees.max_row, min_col=6, max_col=9): 
        for cell in row:
            cell.fill = light_pink_fill

    for col in range(1, len(header) + 1):
        col_letter = get_column_letter(col)
        ws_all_employees.column_dimensions[col_letter].width = 15

    ws_periodic_workers = wb.create_sheet(title="기간제근로자")
    header = ["이름", "직급", "입사일", "부서"]
    ws_periodic_workers.append(header)

    header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") 
    header_font = Font(bold=True)  
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    ) 

    for col_num, header_text in enumerate(header, start=1):
        cell = ws_periodic_workers.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border 

    for row_idx, (_, row) in enumerate(periodic_workers.iterrows(), start=2):
        ws_periodic_workers.append([
            row.get("한글명"), row.get("직급"), row.get("입사일"), row.get("세부부서")
        ])
        for col_idx in range(1, 5):
            ws_periodic_workers.cell(row=row_idx, column=col_idx).border = thin_border

    for col in range(1, 5):
        col_letter = get_column_letter(col)
        ws_periodic_workers.column_dimensions[col_letter].width = 17

    ws_main.append(["구분", "계"] + [None] + position_order) 
    ws_main.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2) 
    ws_main.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2) 
    ws_main.cell(row=1, column=1).value = "구분"
    ws_main.cell(row=1, column=3).value = "계"

    for col in range(1, 4 + len(position_order)): 
        cell = ws_main.cell(row=1, column=col)
        if col > 3: 
            cell.value = position_order[col - 4] 
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

    total_current = sum(department_counts.values)
    total_quota = sum(department_quota.values())
    total_surplus_deficit = total_current - total_quota

    ws_main.append(["정원", total_quota, None] + [position_quota.get(pos, 0) for pos in position_order]) 
    ws_main.append(["현원", total_current, None] + [position_counts.get(pos, 0) for pos in position_order]) 
    ws_main.append(["과부족", total_surplus_deficit, None] + [
        position_counts.get(pos, 0) - position_quota.get(pos, 0) for pos in position_order
    ])

    for row_idx, title in enumerate(["정원", "현원", "과부족"], start=2):
        ws_main.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        ws_main.cell(row=row_idx, column=1).value = title  
        for col in range(1, 5 + len(position_order)):  
            cell = ws_main.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center") 
            cell.font = Font(bold=True) if col == 1 else Font()  
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

    for row_idx, total_value in enumerate([total_quota, total_current, total_surplus_deficit], start=2):
        ws_main.cell(row=row_idx, column=3).value = total_value 
        cell = ws_main.cell(row=row_idx, column=3)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

    current_row = ws_main.max_row + 1

    max_col = ws_main.max_column  
    if max_col > 0: 
        ws_main.delete_cols(max_col)

    for department in department_order:
        total_quota = sum(detailed_quota.get(department, {}).values())
        total_current = sum(grouped_counts.get((department, pos), 0) for pos in position_order)
        total_surplus_deficit = total_current - total_quota

        sub_department_count = 0
        for department_code, department_data in department_hierarchy.items():
            if department_data["name"] == department:
                for sub_code, sub_data in department_data["sub_departments"].items():
                    sub_department_count += 1
                    sub_department_count += len(sub_data["sub_departments"])

        total_merge_rows = 3 + sub_department_count
        ws_main.cell(row=current_row, column=1).value = department
        ws_main.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row + total_merge_rows - 1,
            end_column=1
        )

        ws_main.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_main.cell(row=current_row, column=1).font = Font(bold=True)

        ws_main.cell(row=current_row, column=2).value = "정원"
        ws_main.cell(row=current_row + 1, column=2).value = "현원"
        ws_main.cell(row=current_row + 2, column=2).value = "과부족"

        for row, bg_color, font_color, bold in zip(
            [current_row, current_row + 1, current_row + 2],
            ["FDE9D9", "92D050", "FFFF00"],
            ["000000", "000000", "FF0000"],
            [False, False, True]
        ):
            for col in range(2, 4 + len(position_order)):
                cell = ws_main.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                cell.font = Font(color=font_color, bold=bold)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_main.cell(row=current_row, column=3).value = total_quota
        ws_main.cell(row=current_row + 1, column=3).value = total_current
        ws_main.cell(row=current_row + 2, column=3).value = total_surplus_deficit

        for col_index, position in enumerate(position_order, start=4):
            quota_value = detailed_quota.get(department, {}).get(position, 0)
            current_value = grouped_counts.get((department, position), 0)
            surplus_deficit_value = current_value - quota_value

            ws_main.cell(row=current_row, column=col_index).value = quota_value
            ws_main.cell(row=current_row + 1, column=col_index).value = current_value
            ws_main.cell(row=current_row + 2, column=col_index).value = surplus_deficit_value

        current_row += 3

        for department_code, department_data in department_hierarchy.items():
            if department_data["name"] == department:
                for sub_code, sub_data in department_data["sub_departments"].items():
                    sub_name = sub_data["name"]
                    sub_current = input_df[input_df["부서"].astype(str) == sub_code].shape[0]

                    ws_main.cell(row=current_row, column=2).value = sub_name
                    ws_main.cell(row=current_row, column=3).value = sub_current

                    for col_index, position in enumerate(position_order, start=4):
                        sub_position_current = input_df[
                            (input_df["부서"].astype(str) == sub_code) & (input_df["직급"] == position)
                        ].shape[0]
                        ws_main.cell(row=current_row, column=col_index).value = sub_position_current

                    current_row += 1

                    for sub_sub_code, sub_sub_data in sub_data["sub_departments"].items():
                        sub_sub_name = sub_sub_data["name"]
                        sub_sub_current = input_df[input_df["부서"].astype(str) == sub_sub_code].shape[0]

                        ws_main.cell(row=current_row, column=2).value = sub_sub_name
                        ws_main.cell(row=current_row, column=3).value = sub_sub_current

                        for col_index, position in enumerate(position_order, start=4):
                            sub_sub_position_current = input_df[
                                (input_df["부서"].astype(str) == sub_sub_code) & (input_df["직급"] == position)
                            ].shape[0]
                            ws_main.cell(row=current_row, column=col_index).value = sub_sub_position_current

                        current_row += 1

        ws_main.column_dimensions['B'].width = 18

    department_names = [dept for dept, _ in sorted(department_counts.items(), key=lambda x: department_order.index(x[0]))]
    current_values = [department_counts[dept] for dept in department_names]
    quota_values = [department_quota.get(dept, 0) for dept in department_names]
    surplus_deficit = [current - quota for current, quota in zip(current_values, quota_values)] 

    plt.figure(figsize=(12, 8))
    x = range(len(department_names))
    width = 0.3 

    bars_current = plt.bar(x, current_values, width=width, label="현원", align="center", color='skyblue')
    bars_quota = plt.bar([i + width for i in x], quota_values, width=width, label="정원", align="center", color='lightgreen')
    bars_surplus = plt.bar([i + 2 * width for i in x], surplus_deficit, width=width, label="과부족", align="center", color='salmon')

    plt.xticks([i + width for i in x], department_names, rotation=45)

    for bars, data in zip([bars_current, bars_quota, bars_surplus], [current_values, quota_values, surplus_deficit]):
        for bar, value in zip(bars, data):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5, str(value), ha="center", va="bottom", fontsize=10)

    plt.title("중구시설관리공단 인사 현황판", fontsize=16)
    plt.xlabel("부서", fontsize=11)
    plt.ylabel("인원", fontsize=11)
    plt.legend(fontsize=11)
    plt.tight_layout()

    def show_graph_and_preview_interface(graphs, excel_file):
        """그래프와 엑셀 미리보기 인터페이스"""
        root = Toplevel()
        root.title("그래프 및 인원 상세보기")
        root.geometry("1400x1000")

        def on_close():
            root.destroy()
            sys.exit()

        root.protocol("WM_DELETE_WINDOW", on_close)

        graph_frame = Frame(root, width=1200, height=700, bg="white")
        graph_frame.pack(side="top", fill="both", expand=True)

        figure_canvas = None
        current_index = [0]

        def update_canvas():
            nonlocal figure_canvas
            if figure_canvas:
                figure_canvas.get_tk_widget().pack_forget()
            fig = graphs[current_index[0]]
            figure_canvas = FigureCanvasTkAgg(fig, master=graph_frame)
            figure_canvas.draw()
            figure_canvas.get_tk_widget().pack(fill="both", expand=True)

        def next_graph():
            current_index[0] = (current_index[0] + 1) % len(graphs)
            update_canvas()

        def prev_graph():
            current_index[0] = (current_index[0] - 1) % len(graphs)
            update_canvas()

        def save_graph():
            fig = graphs[current_index[0]]
            file_name = f"graph_{current_index[0] + 1}.png"
            fig.savefig(file_name, bbox_inches="tight")
            print(f"✅ 그래프가 저장되었습니다: {file_name}")

            button_frame = Frame(root, height=50, bg="lightgray")
            button_frame.pack(side="top", fill="x")
        
        def preview_excel():
            """엑셀 내용을 미리보기로 표시"""
            preview_window = Toplevel(root)
            preview_window.title("인원 상세 정보")
            preview_window.geometry("800x600")

            style = Style(preview_window)
            style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
            style.configure("Treeview", rowheight=25, font=("Arial", 10))

            df = pd.read_excel(excel_file)
            df = df.fillna("")

            tree = Treeview(preview_window, columns=list(df.columns), show="headings", height=20)

            for col in df.columns:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor="center")

            tree.tag_configure("quota", background="#FDE9D9", foreground="black")  
            tree.tag_configure("current", background="#92D050", foreground="black")  
            tree.tag_configure("surplus_deficit", background="#FFFF00", foreground="red", font=("Arial", 10, "bold"))
            tree.tag_configure("department", background="#FFFFFF", foreground="black")  
            tree.tag_configure("default", background="#FFFFFF", foreground="black")  

            for i, (_, row) in enumerate(df.iterrows()):
                values = list(row)
                tag = "default"  

                if values[0] == "정원" or values[1] == "정원":
                    tag = "quota"
                elif values[0] == "현원" or values[1] == "현원":
                    tag = "current"
                elif values[0] == "과부족" or values[1] == "과부족":
                    tag = "surplus_deficit"
                elif values[0] in department_counts.index:  
                    tag = "department"

                tree.insert("", END, values=values, tags=(tag,))

            vsb = Scrollbar(preview_window, orient="vertical", command=tree.yview)
            vsb.pack(side="right", fill="y")
            tree.configure(yscrollcommand=vsb.set)

            tree.pack(fill="both", expand=True)

            def show_name_list(event):
                """선택된 부서/세부부서/직급 별 인원 목록 표시"""
                selected_item = tree.focus()
                selected_values = tree.item(selected_item, "values")

                if not selected_values:
                    print("⚠️ 선택된 항목에 값이 없습니다.")
                    return

                selected_department = selected_values[0].strip() if len(selected_values) > 0 else ""
                selected_sub_department = selected_values[1].strip() if len(selected_values) > 1 else ""
                selected_position = selected_values[2].strip() if len(selected_values) > 2 else ""

                print(f"🟢 선택한 값: 부서 = '{selected_department}', 세부부서 = '{selected_sub_department}', 직급 = '{selected_position}'")

                if selected_department == "":
                    matched_row = input_df[input_df["세부부서"] == selected_sub_department]
                    if not matched_row.empty:
                        selected_department = matched_row.iloc[0]["부서.1"]

                if selected_sub_department in ["정원", "현원", "과부족"]:
                    print(f"⚠️ '{selected_sub_department}'은(는) 통계 행이므로 필터링 제외")
                    return

                input_df["부서.1"] = input_df["부서.1"].astype(str).str.strip().replace({"nan": "", "NaN": "", None: ""})
                input_df["세부부서"] = input_df["세부부서"].astype(str).str.strip().replace({"nan": "", "NaN": "", None: ""})
                input_df["직급"] = input_df["직급"].astype(str).str.strip().replace({"nan": "", "NaN": "", None: ""})

                selected_department = selected_department.strip()
                selected_sub_department = selected_sub_department.strip()
                selected_position = selected_position.strip()

                department_filtered_df = input_df.copy()

                if selected_department in input_df["부서.1"].values:
                    department_filtered_df = department_filtered_df[
                        department_filtered_df["부서.1"] == selected_department
                    ]

                if selected_sub_department in input_df["세부부서"].values:
                    department_filtered_df = department_filtered_df[
                        department_filtered_df["세부부서"] == selected_sub_department
                    ]

                if selected_position in input_df["직급"].values:
                    department_filtered_df = department_filtered_df[
                        department_filtered_df["직급"] == selected_position
                    ]

                print(f"🔍 필터링된 데이터 개수: {len(department_filtered_df)}")

                if department_filtered_df.empty:
                    print(f"⚠️ '{selected_department} - {selected_sub_department} - {selected_position}'에 해당하는 데이터가 없습니다.")
                    return

                name_list = department_filtered_df["한글명"].tolist()

                name_window = Toplevel(preview_window)
                name_window.title(f"{selected_department} - 이름 목록")
                name_window.geometry("400x300")

                Label(name_window, text=f"{selected_department} {selected_sub_department} {selected_position} - {len(name_list)}명", font=("Arial", 12)).pack(pady=10)

                name_listbox = Listbox(name_window)
                for name in name_list:
                    name_listbox.insert(END, name)

                name_listbox.pack(fill="both", expand=True, padx=10, pady=10)

                name_listbox.bind("<Double-1>", lambda event: show_employee_details(event, name_listbox))
            def show_employee_details(event, listbox):
                """선택된 직원의 상세 정보를 보여주는 함수"""
                try:
                    selected_index = listbox.curselection()
                    if not selected_index:
                        print("⚠️ 선택된 항목이 없습니다.")
                        return

                    selected_name = listbox.get(selected_index[0]).strip()

                    employee_row = input_df[input_df["한글명"] == selected_name]

                    if employee_row.empty:
                        print(f"⚠️ '{selected_name}'에 대한 정보가 없습니다.")
                        return

                    employee_info = employee_row.iloc[0]

                    details_window = Toplevel(listbox.master)
                    details_window.title(f"{selected_name} 상세 정보")
                    details_window.geometry("400x400")

                    Label(details_window, text=f"이름: {employee_info['한글명']}", font=("Arial", 12, "bold")).pack(pady=5)
                    Label(details_window, text=f"사원번호: {employee_info['사원번호']}", font=("Arial", 12)).pack(pady=5)
                    Label(details_window, text=f"직급: {employee_info['직급']}", font=("Arial", 12)).pack(pady=5)
                    Label(details_window, text=f"입사일: {employee_info['입사일']}", font=("Arial", 12)).pack(pady=5)
                    Label(details_window, text=f"부서: {employee_info['세부부서']}", font=("Arial", 12)).pack(pady=5)
                    Label(details_window, text=f"생년월일: {employee_info['생년월일']}", font=("Arial", 12)).pack(pady=5)

                    if "임금피크(1)" in employee_info and not pd.isna(employee_info["임금피크(1)"]):
                        Label(details_window, text=f"임금피크(1): {employee_info['임금피크(1)']}", font=("Arial", 12)).pack(pady=5)
                    if "임금피크(2)" in employee_info and not pd.isna(employee_info["임금피크(2)"]):
                        Label(details_window, text=f"임금피크(2): {employee_info['임금피크(2)']}", font=("Arial", 12)).pack(pady=5)
                    if "임금피크(3)" in employee_info and not pd.isna(employee_info["임금피크(3)"]):
                        Label(details_window, text=f"임금피크(3): {employee_info['임금피크(3)']}", font=("Arial", 12)).pack(pady=5)
                    if "정년" in employee_info and not pd.isna(employee_info["정년"]):
                        Label(details_window, text=f"정년: {employee_info['정년']}", font=("Arial", 12)).pack(pady=5)

                    Button(details_window, text="닫기", command=details_window.destroy).pack(pady=10)

                except Exception as e:
                    print(f"⚠️ 직원 정보 창 표시 중 오류 발생: {e}")

            tree.bind("<Double-1>", show_name_list)  

        button_frame = Frame(root, height=50, bg="lightgray")
        button_frame.pack(side="top", fill="both", expand=True)

        Button(button_frame, text="이전 그래프", command=prev_graph, height=2, width=15).pack(side="left", padx=10, pady=5)
        Button(button_frame, text="다음 그래프", command=next_graph, height=2, width=15).pack(side="left", padx=10, pady=5)
        Button(button_frame, text="그래프 저장", command=save_graph, height=2, width=15).pack(side="left", padx=10, pady=5)
        Button(button_frame, text="부서 별 인원 상세보기", command=preview_excel, height=2, width=19).pack(side="right", padx=10, pady=5)

        update_canvas()
        root.mainloop()

    sorted_indices = [department_names.index(name) for name in department_order]

    department_names = [department_names[i] for i in sorted_indices]
    current_values = [current_values[i] for i in sorted_indices]
    quota_values = [quota_values[i] for i in sorted_indices]
    surplus_deficit = [surplus_deficit[i] for i in sorted_indices]

    def plot_graphs(department_names, current_values, quota_values, surplus_deficit):
        """여러 그래프를 생성하여 저장"""
        graphs = []

        sorted_indices = [department_names.index(name) for name in department_order]
        department_names = [department_names[i] for i in sorted_indices]
        current_values = [current_values[i] for i in sorted_indices]
        quota_values = [quota_values[i] for i in sorted_indices]
        surplus_deficit = [surplus_deficit[i] for i in sorted_indices]

        fig1 = plt.figure(figsize=(15, 9)) 
        x = range(len(department_names))
        width = 0.3

        bars_current = plt.bar(x, current_values, width=width, label="현원", align="center", color='skyblue')
        bars_quota = plt.bar([i + width for i in x], quota_values, width=width, label="정원", align="center", color='lightgreen')
        bars_surplus = plt.bar([i + 2 * width for i in x], surplus_deficit, width=width, label="과부족", align="center", color='salmon')

        plt.xticks([i + width for i in x], department_names, rotation=45)
        for bars, data in zip([bars_current, bars_quota, bars_surplus], [current_values, quota_values, surplus_deficit]):
            for bar, value in zip(bars, data):
                plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5, str(value), ha="center", va="bottom", fontsize=10)

        plt.title("부서별 현원/정원/과부족 비교", fontsize=16, pad=20) 
        plt.xlabel("부서", fontsize=11)
        plt.ylabel("인원", fontsize=11)
        plt.legend(fontsize=11)
        plt.tight_layout() 
        plt.subplots_adjust(top=0.90, left=0.1, right=0.9) 
        graphs.append(fig1)

        fig2 = plt.figure(figsize=(15, 9))
        deficit_ratio = [round(s / q * 100, 2) if q != 0 else 0 for s, q in zip(surplus_deficit, quota_values)]
        plt.bar(department_names, deficit_ratio, color="salmon")
        plt.title("부서별 과부족 비율", fontsize=16, pad=20)
        plt.xlabel("부서", fontsize=11)
        plt.subplots_adjust(top=0.90, left=0.1, right=0.9)
        plt.ylabel("과부족 비율 (%)", fontsize=11)
        for i, v in enumerate(deficit_ratio):
            plt.text(i, v + 0.5, f"{v}%", ha="center", fontsize=10)
        plt.tight_layout() 
        plt.subplots_adjust(top=0.85, left=0.1, right=0.9) 
        graphs.append(fig2)

        return graphs

    graphs = plot_graphs(department_names, current_values, quota_values, surplus_deficit)

    ws_hierarchy = wb.create_sheet(title="부서별 세부 인원")
    ws_hierarchy.append(["부서", "현원", "이름 목록"])

    ws_hierarchy.column_dimensions['A'].width = 18 
    ws_main.column_dimensions['A'].width = 13


    updated_hierarchy = traverse_hierarchy_and_count(department_hierarchy, input_df)
    write_hierarchy_to_excel(ws_hierarchy, updated_hierarchy, input_df)

    for ws in [ws_main, ws_hierarchy]:
        max_col = ws.max_column 
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    file_name = f"정원_현황_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(file_name)
    print(f"✅ 엑셀 파일이 생성되었습니다: {file_name}")

    show_graph_and_preview_interface(graphs, file_name)

def main(skip_misclassified_check=False):
    print("✅ 사원 정보가 적힌 엑셀 파일을 첨부해주세요.")
    file_path = select_file()

    if not file_path:
        print("파일이 선택되지 않았습니다.")
        return

    input_df, periodic_workers = preprocess_data(pd.read_excel(file_path))

    if "세부부서" not in input_df.columns:
        print("세부부서 열이 없습니다. 데이터 전처리를 확인하세요.")
        return

    if not skip_misclassified_check:
        misclassified_df = detect_misclassified_employees(input_df)

        if not misclassified_df.empty:
            show_misclassified_employees_ui(misclassified_df, input_df, file_path, periodic_workers)

            main(skip_misclassified_check=True)
            return  

    save_to_excel_and_continue(input_df, periodic_workers, file_path)

    department_counts = input_df["부서.1"].value_counts()
    position_counts = input_df["직급"].value_counts()
    position_counts["2급"] = 0
    position_counts["비서"] = 0

    positions = position_counts.index.tolist()
    positions.extend(["2급", "비서"])

    grouped_counts = input_df.groupby(["부서.1", "직급"]).size()
    sub_department_counts = input_df["세부부서"].value_counts()

    total_quota = 100  
    departments = department_counts.index.tolist()
    positions = position_counts.index.tolist()

    use_custom_order, custom_order = get_sorting_order(departments, positions)

    if use_custom_order:
        departments = custom_order["department"]
        positions = custom_order["position"]

    detailed_quota = get_quota_input(departments, positions)

    department_hierarchy = {
     "85": {"name": "임원", "sub_departments": {
        "8500": {"name": "임원", "sub_departments": {}}
     }},
    "84": {
        "name": "안전감사실",
        "sub_departments": {
            "8400": {"name": "안전감사실_부장", "sub_departments": {}},
            "8401": {"name": "청렴감사팀", "sub_departments": {}},
            "8402": {"name": "안전보건팀", "sub_departments": {}}
        }
    },
    "81": {
        "name": "전략기획실",
        "sub_departments": {
            "8100": {"name": "전략기획부_부장", "sub_departments": {}},
            "8101": {"name": "기획조정팀", "sub_departments": {}},
            "8103": {"name": "소통경영팀", "sub_departments": {}}
        }
    },
    "82": {
        "name": "경영지원부",
        "sub_departments": {
            "8200": {"name": "경영지원부_부장", "sub_departments": {}},
            "8201": {"name": "인사총무팀", "sub_departments": {}},
            "8202": {"name": "재정지원팀", "sub_departments": {}}
        }
    },
    "83": {
        "name": "체육사업부",
        "sub_departments": {
            "8300": {"name": "체육사업부_부장", "sub_departments": {}},
            "8301": {
                "name": "체육사업1팀",
                "sub_departments": {
                    "8352": {"name": "충무", "sub_departments": {}},
                    "8351": {"name": "무학봉", "sub_departments": {}},
                    "8358": {"name": "훈련원공원", "sub_departments": {}}
                }
            },
            "8302": {
                "name": "체육사업2팀",
                "sub_departments": {
                    "8354": {"name": "회현", "sub_departments": {}},
                    "8353": {"name": "손기정", "sub_departments": {}},
                    "8361": {"name": "남산", "sub_departments": {}},
                    "8360": {"name": "장충", "sub_departments": {}}
                }
            }
        }
    },
    "86": {
        "name": "주차사업부",
        "sub_departments": {
            "8600": {"name": "주차사업부_부장", "sub_departments": {}},
            "8601": {"name": "주차사업1팀", "sub_departments": {}},
            "8602": {"name": "주차사업2팀", "sub_departments": {}}
        }
    },
    "91": {
        "name": "시설관리부",
        "sub_departments": {
            "9100": {"name": "시설관리부_부장", "sub_departments": {}},
            "9101": {"name": "시설관리1팀", "sub_departments": {}},
            "9102": {
                "name": "시설관리2팀",
                "sub_departments": {
                    "9151": {"name": "중립종합복지센터", "sub_departments": {}},
                    "9152": {"name": "중구종합복지센터", "sub_departments": {}},
                    "9153": {"name": "신당누리센터", "sub_departments": {}},
                    "9154": {"name": "교육지원센터", "sub_departments": {}}
                }
            },
            "9103": {
                "name": "공공시설팀",
                "sub_departments": {
                    "9155": {"name": "중구구민회관", "sub_departments": {}},
                    "9156": {"name": "충무창업큐브", "sub_departments": {}},
                    "9157": {"name": "공중공원화장실", "sub_departments": {}}
                }
            }
        }
    },
    "89": {"name": "사회서비스단", 
    "sub_departments": {
        "8900": {"name": "사회서비스단",  "sub_departments": {}}
    }}
}

    department_counts = pd.Series({dept: department_counts[dept] for dept in departments if dept in department_counts})
    position_counts = pd.Series({pos: position_counts[pos] for pos in positions if pos in position_counts})

    create_excel_file(
        department_counts,
        position_counts,
        grouped_counts,
        total_quota,
        {}, 
        {}, 
        detailed_quota,
        sub_department_counts,
        department_hierarchy,
        input_df,
        departments, 
        positions,
        periodic_workers 
    )

if __name__ == "__main__":
    main()